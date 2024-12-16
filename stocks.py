from nsetools import Nse
import openpyxl
from pathlib import Path

def fetch_nse_prices(scrips):
    """
    Fetches current market prices for a list of scrip symbols from NSE.
    
    Args:
        scrips (list): List of NSE scrip symbols.
    
    Returns:
        dict: A dictionary with scrip symbols as keys and their current prices as values.
    """
    nse = Nse()
    prices = {}
    for scrip in scrips:
        try:
            data = nse.get_quote(scrip)
            if data:
                prices[scrip] = data['lastPrice']
        except Exception as e:
            print(f"Error fetching data for {scrip}: {e}")
    return prices

def update_excel(file_path, column, start_row, stock_name_to_scrip, prices):
    """
    Updates the specified column in an Excel file with the latest prices.
    
    Args:
        file_path (str): Path to the Excel file.
        column (str): Column letter to update (e.g., 'F').
        start_row (int): Row number to start updating from (e.g., 2).
        stock_name_to_scrip (dict): Mapping of stock names to NSE scrip symbols.
        prices (dict): Latest prices fetched from NSE.
    """
    try:
        # Load the Excel workbook
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Update the column with prices
        for i, row in enumerate(sheet.iter_rows(min_row=start_row), start=start_row):
            stock_name = row[0].value  # Assuming stock names are in column A
            if stock_name in stock_name_to_scrip:
                scrip = stock_name_to_scrip[stock_name]
                if scrip in prices:
                    sheet[f"{column}{i}"] = prices[scrip]

        # Save the updated workbook
        workbook.save(file_path)
        print(f"Successfully updated {file_path}")
    except Exception as e:
        print(f"Error updating Excel file: {e}")

if __name__ == "__main__":
    # Define the path to your Excel file and relevant details
    file_path = Path("/home/aniketdatar/Downloads/Final_VVD_P&L")  # Replace with your file path
    column_to_update = "F"
    start_row = 2

    # Define a mapping of stock names to their NSE scrip symbols
    stock_name_to_scrip = {
        "ASIAN PAINTS": "ASIANPAINT",
        "BRITANNIA INDUSTRIES": "BRITANNIA",
        "HAPPIEST MINDS TECH": "HAPPSTMNDS",
        "HCL TECHNOLOGIES": "HCLTECH",
        "ITC": "ITC",
        "MAHINDRA & MAHINDRA": "M&M",
        "PTC INDIA": "PTC",
        "TATA CHEMICALS": "TATACHEM",
        "TATA ELXSI": "TATAELXSI",
        "TATA POWER": "TATAPOWER",
        "TATA STEEL": "TATASTEEL",
        "INFOSYS": "INFY",
        "WIPRO": "WIPRO",
        "ADANI PORTS": "ADANIPORTS",
        "DELTACORP": "DELTACORP",
        "DRREDDY": "DRREDDY",
        "GRASIM": "GRASIM",
        "HAVELLS": "HAVELLS",
        "INDIAN HOTELS": "INDHOTEL",
        "SIEMENS": "SIEMENS",
        "IRCTC": "IRCTC",
        "STATE BANK OF INDIA": "SBIN",
        "TRENT": "TRENT",
        "LIC INDIA": "LICI",
        "TCS": "TCS"
        #"RelianceIndustries": "RELIANCE",
        #"HDFC Bank": "HDFCBANK"
    }

    # Extract stock names from column A of the Excel file
    stock_names = []
    if file_path.exists():
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        stock_names = [sheet[f"A{row}"].value for row in range(start_row, sheet.max_row + 1)]

    # Fetch prices for the corresponding scrips
    scrips_to_fetch = [stock_name_to_scrip[name] for name in stock_names if name in stock_name_to_scrip]
    print("Fetching live data from NSE...")
    stock_prices = fetch_nse_prices(scrips_to_fetch)
    print("Fetched Prices:", stock_prices)

    # Update the Excel file with the latest prices
    update_excel(file_path, column_to_update, start_row, stock_name_to_scrip, stock_prices)
