from yahooquery import Ticker
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string
import openpyxl
from pathlib import Path
from datetime import datetime

# Your stock tickers dictionary
stock_name_to_scrip = {
    "ASIAN PAINTS": "ASIANPAINT.NS",
    "BRITANNIA INDUSTRIES": "BRITANNIA.NS",
    "HAPPIEST MINDS TECH": "HAPPSTMNDS.NS",
    "HCL TECHNOLOGIES": "HCLTECH.NS",
    "ITC": "ITC.NS",
    "MAHINDRA & MAHINDRA": "M&M.NS",
    "PTC INDIA": "PTC.NS",
    "TATA CHEMICALS": "TATACHEM.NS",
    "TATA ELXSI": "TATAELXSI.NS",
    "TATA POWER": "TATAPOWER.NS",
    "TATA STEEL": "TATASTEEL.NS",
    "INFOSYS": "INFY.NS",
    "WIPRO": "WIPRO.NS",
    "ADANI PORTS": "ADANIPORTS.NS",
    "DRREDDY": "DRREDDY.NS",
    "GRASIM": "GRASIM.NS",
    "HAVELLS": "HAVELLS.NS",
    "INDIAN HOTELS": "INDHOTEL.NS",
    "SIEMENS": "SIEMENS.NS",
    "ENRIN": "ENRIN.NS",
    "IRCTC": "IRCTC.NS",
    "STATE BANK OF INDIA": "SBIN.NS",
    "TRENT": "TRENT.NS",
    "BAJAJ FINANCE": "BAJFINANCE.NS",
    "INDUSIND BANK": "INDUSINDBK.NS",
    "ABFRL": "ABFRL.NS",
    "ABLBL": "ABLBL.NS",
    "TEJASNET": "TEJASNET.NS",
    "HYUNDAI": "HYUNDAI.NS",
    "LIC INDIA": "LICI.NS",
    "TCS": "TCS.NS",
}

file_path = Path(r"C:\Users\Admin\Downloads\Stock_Price_Automation\Final_P&L_auto.xlsx")
column_to_update = "F"
start_row = 2

def fetch_prices(stock_dict):
    """
    Fetch live CMP if available, else fallback to previous close.
    Return dict: symbol -> price (rounded to 2 decimals) or "N/A".
    """
    symbols = list(stock_dict.values())
    tickers = Ticker(symbols)
    price_data = tickers.price

    prices = {}
    for name, symbol in stock_dict.items():
        info = price_data.get(symbol)
        price = None
        if info and isinstance(info, dict):
            # Try live price first
            price = info.get("regularMarketPrice")
            # Fallback to previous close
            if price is None:
                price = info.get("regularMarketPreviousClose")
        if price is None:
            prices[symbol] = "N/A"
        else:
            prices[symbol] = round(price, 2)
    return prices

def update_excel(file_path, column, start_row, stock_dict, prices):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Update header with current date
        current_date = datetime.now().strftime("CMP ON %b %d %Y")
        sheet[f"{column}1"] = current_date

        # Update prices in the Excel
        for row in range(start_row, sheet.max_row + 1):
            stock_name = sheet[f"A{row}"].value
            if stock_name in stock_dict:
                symbol = stock_dict[stock_name]
                price = prices.get(symbol, "N/A")
                sheet[f"{column}{row}"] = price

        # Apply color formatting to column K (Green if >100, Red if <0)
        col_k_idx = column_index_from_string("K")
        for row in range(start_row, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col_k_idx)
            if isinstance(cell.value, (int, float)):
                if cell.value > 100:
                    cell.font = Font(color="00FF00")  # Green
                elif cell.value < 0:
                    cell.font = Font(color="FF0000")  # Red

        workbook.save(file_path)
        print(f"Excel updated successfully at {file_path}")

    except Exception as e:
        print(f"Error updating Excel: {e}")

if __name__ == "__main__":
    if file_path.exists():
        print("Fetching live prices...")
        prices = fetch_prices(stock_name_to_scrip)
        print("Prices fetched:", prices)
        update_excel(file_path, column_to_update, start_row, stock_name_to_scrip, prices)
    else:
        print(f"Excel file not found: {file_path}")
